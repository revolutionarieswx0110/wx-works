import sys,os  #Demo9_13.py
from PyQt5.QtWidgets import (QApplication,QWidget,QMenuBar,QPlainTextEdit,QAbstractItemView,QStatusBar,
                            QHBoxLayout,QVBoxLayout,QListView,QLabel,QPushButton,QFileDialog,QToolBar)
from PyQt5.QtGui import QIcon,QFont
from PyQt5.QtCore import QStringListModel,QModelIndex,Qt
from openpyxl import load_workbook
import jieba
jieba.set_dictionary("./dict.txt")
jieba.initialize()
import jieba.analyse
jieba.analyse.set_idf_path("./idf.txt")
import jieba.posseg
jieba.set_dictionary("./dict.txt")
jieba.initialize()
import chardet
sheet_name="sheet10"
choose_data=[]
def get_encoding(file):#获取文本编码方式
	with open(file,'rb') as f:
		return chardet.detect(f.read())['encoding']

def dosegment_all(sentence):
    '''
    带词性标注，对句子进行分词，不排除停词等
    :param sentence:输入字符
    :return:
    '''
    global  sentence_seged
    sentence_seged = jieba.posseg.cut(sentence.strip())
    outstr = ''
    for x in sentence_seged:
        outstr += "{}/{},".format(x.word, x.flag)
    # 上面的for循环可以用python递推式构造生成器完成
    # outstr = ",".join([("%s/%s" %(x.word,x.flag)) for x in sentence_seged])
    return outstr

#数字
def qm():
    excludes={}  #构建无意义词语集合
    counts={}    #构建空字典，存放词语和词频
    for word in words:
        if  "m"  not in word:           #跳过标点符号
            continue
        else:
            word=word.replace("/m","")#跳过非ren名词
            rword=word
        counts[rword]=counts.get(rword,0)+1        #统计词语出现的次数
        #删除无意义的词语
    for word in excludes:
        del counts[word]
        #按词语出现次数排序
    items=list(counts.items())
    items.sort(key=lambda x:x[1],reverse=True)
    global night_data
    night_data=items
#其它
def qt():
    excludes={}  #构建无意义词语集合
    counts={}    #构建空字典，存放词语和词频
    for word in words:
        if "v" in word or "a" in word or"n" in word:           #跳过标点符号
            continue
        else:
            lkd= word.split("/")
            lkd.pop()
            word="".join(lkd)
            #跳过非ren名词
            rword=word
        counts[rword]=counts.get(rword,0)+1        #统计词语出现的次数
        #删除无意义的词语
    for word in excludes:
        del counts[word]
        #按词语出现次数排序
    items=list(counts.items())
    items.sort(key=lambda x:x[1],reverse=True)
    global eighth_data
    eighth_data=items
#动名词
def qvn():
    excludes={}  #构建无意义词语集合
    counts={}    #构建空字典，存放词语和词频
    for word in words:
        if  "vn"  not in word:           #跳过标点符号
            continue
        else:
            word=word.replace("/vn","")#跳过非ren名词
            rword=word
        counts[rword]=counts.get(rword,0)+1        #统计词语出现的次数
        #删除无意义的词语
    for word in excludes:
        del counts[word]
        #按词语出现次数排序
    items=list(counts.items())
    items.sort(key=lambda x:x[1],reverse=True)
    global seventh_data
    seventh_data=items
#动词
def qv():
    excludes={}  #构建无意义词语集合
    counts={}    #构建空字典，存放词语和词频
    for word in words:
        if  "v"  not in word:           #跳过标点符号
            continue
        elif "vn" in word or "uv" in word:
            continue
        elif "vg" in word:
            word = word.replace("/vg","")  # 跳过非ren名词
            rword = word
        else:
            word=word.replace("/v","")#跳过非ren名词
            rword=word
        counts[rword]=counts.get(rword,0)+1        #统计词语出现的次数
        #删除无意义的词语
    for word in excludes:
        del counts[word]
        #按词语出现次数排序
    items=list(counts.items())
    items.sort(key=lambda x:x[1],reverse=True)
    global sixth_data
    sixth_data=items
#形容词
def qa():
    excludes={}  #构建无意义词语集合
    counts={}    #构建空字典，存放词语和词频
    for word in words:
        if "a"  not in word:           #跳过标点符号
            continue
        elif "an" in word:
            word = word.replace("/an", "")  # 跳过非ren名词
            rword = word
        elif "ad" in word:
            word = word.replace("/ad", "")  # 跳过非ren名词
            rword = word
        elif "ag" in word:
            word = word.replace("/ag", "")  # 跳过非ren名词
            rword = word
        else:
            word=word.replace("/a","")#跳过非ren名词
            rword=word
        counts[rword]=counts.get(rword,0)+1        #统计词语出现的次数
        #删除无意义的词语
    for word in excludes:
        del counts[word]
        #按词语出现次数排序
    items=list(counts.items())
    items.sort(key=lambda x:x[1],reverse=True)
    global fifth_data
    fifth_data=items
#名词
def qn():
    excludes = {}  # 构建无意义词语集合
    counts = {}  #vn 构建空字典，存放词语和词频
    for word in words:
        if "n" not in word:  # 跳过标点符号
            continue
        elif "nr" in word or  "an" in word or "vn" in word or "vr" in word or "ns" in word or "nt" in word or "nz" in word or   "ng" in word:
            continue
        else:
            word = word.replace("/n", "")  # 跳过非ren名词
            rword = word
        counts[rword] = counts.get(rword, 0) + 1  # 统计词语出现的次数
        # 删除无意义的词语
    for word in excludes:
        del counts[word]
        # 按词语出现次数排序
    items = list(counts.items())
    items.sort(key=lambda x: x[1], reverse=True)
    global fourth_data
    fourth_data = items
#机构名
def qnt():
    excludes={}  #构建无意义词语集合
    counts={}    #构建空字典，存放词语和词频
    for word in words:
        if  "nt"  not in word:           #跳过标点符号
            continue
        else:
            word=word.replace("/nt","")#跳过非ren名词
            rword=word
        counts[rword]=counts.get(rword,0)+1        #统计词语出现的次数
        #删除无意义的词语
    for word in excludes:
        del counts[word]
        #按词语出现次数排序
    items=list(counts.items())
    items.sort(key=lambda x:x[1],reverse=True)
    global third_data
    third_data=items
#地名
def qns():
    excludes={}  #构建无意义词语集合
    counts={}    #构建空字典，存放词语和词频
    for word in words:
        if "ns"  not in word:           #跳过标点符号
            continue
        else:
            word=word.replace("/ns","")#跳过非ren名词
            rword=word
        counts[rword]=counts.get(rword,0)+1        #统计词语出现的次数
        #删除无意义的词语
    for word in excludes:
        del counts[word]
        #按词语出现次数排序
    items=list(counts.items())
    items.sort(key=lambda x:x[1],reverse=True)
    global second_data
    second_data=items
#人名
def qnr():
    excludes={}  #构建无意义词语集合
    counts={}    #构建空字典，存放词语和词频
    for word in words:
        if "nr"  not in word:           #跳过标点符号
            continue
        elif "nrfg" in word:
            word = word.replace("/nrfg", "")  # 跳过非ren名词
            rword = word
        elif "nrt" in word:
            word = word.replace("/nrt", "")  # 跳过非ren名词
            rword = word
        else:
            word=word.replace("/nr","")#跳过非ren名词
            rword=word
        counts[rword]=counts.get(rword,0)+1        #统计词语出现的次数
        #删除无意义的词语
    for word in excludes:
        del counts[word]
        #按词语出现次数排序
    items=list(counts.items())
    items.sort(key=lambda x:x[1],reverse=True)
    global first_data
    first_data=items

i=""
count_w=0
counts={}
class myWindow(QWidget):
    def __init__(self,parent=None):
        super().__init__(parent)
        self.showMaximized()
        self.setWindowTitle("提取中文关键词")
        self.fileDialog = QFileDialog(self)
        self.widget_setupUi()

    def widget_setupUi(self):  #建立主程序界面
        self.setStyleSheet("QWidget{background-color:rgb(250,230,210)}")
        menuBar = QMenuBar(self)#定义菜单栏
        menuBar.setStyleSheet("background-color:rgb(212,250,248)")
        file_menu = menuBar.addMenu("文件(&F)")  #定义文件菜单
        file_menu.setStyleSheet("background-color:rgb(212,232,223)")
        file_menu.addSeparator()
        action_open = file_menu.addAction("打开")
        action_choose = file_menu.addAction("选择")
        action_saveAs = file_menu.addAction("保存处理文件至")
        action_help = file_menu.addAction("帮助")

        self.plainText = QPlainTextEdit(self)  #显示数据控件
        self.plainText.setStyleSheet("QPlainTextEdit{background-color:rgb(210,212,230)}")
        self.plainText.setReadOnly(True)
        font = QFont()
        font.setFamily("SimSun")  # 宋体
        font.setBold(True)
        font.setPointSizeF(16)
        self.plainText.setFont(font)
        self.plainText.appendPlainText("贵州师范大学附属中学研究性学习0202小组")
        v= QVBoxLayout(self)  #主程序界面的布局
        v.addWidget(menuBar)
        v.addWidget(self.plainText)

        action_open.triggered.connect(self.action_open_triggered)
        action_saveAs.triggered.connect(self.action_saveAs_triggered)
        action_choose.triggered.connect(self.action_choose_triggered)
        action_help.triggered.connect(self.action_help_triggered)

        action_open.setShortcut('ctrl+O')
        action_choose.setShortcut('ctrl+W')
        action_saveAs.setShortcut('ctrl+S')
        action_help.setShortcut('ctrl+H')

    def action_open_triggered(self): #打开txt文件，获取文件名称
        n=""
        self.fileDialog.setAcceptMode(QFileDialog.AcceptOpen)
        self.fileDialog.setFileMode(QFileDialog.ExistingFile)
        self.fileDialog.setNameFilter("文本文件(*.txt)")
        self.plainText.appendPlainText("正在处理文本，请稍等")
        if self.fileDialog.exec():
            n=self.fileDialog.selectedFiles()[0]
            self.plainText.appendPlainText(n)
            global r
            global words
            r = n
            txt = open(r, "r", encoding=get_encoding(r)).read()
            global count_w
            count_w=len(txt)
            words = dosegment_all(txt)
            words = words.split(',')
            qnr()
            qns()
            qn()
            qnt()
            qv()
            qvn()
            qa()
            qt()
            qm()
            #数据转移
            import openpyxl
            # 生成一个 Workbook 的实例化对象，wb即代表一个工作簿（一个 Excel 文件）
            wb = openpyxl.Workbook()
            # 获取活跃的工作表，ws代表wb(工作簿)的一个工作表
            ws = wb.active
            # 更改工作表ws的title
            ws.title = 'sheet1'
            ws2 = wb.create_sheet("sheet2")
            ws3 = wb.create_sheet("sheet3")
            ws4 = wb.create_sheet("sheet4")
            ws5 = wb.create_sheet("sheet5")
            ws6 = wb.create_sheet("sheet6")
            ws7 = wb.create_sheet("sheet7")
            ws8 = wb.create_sheet("sheet8")
            ws9 = wb.create_sheet("sheet9")
            ws10 = wb.create_sheet("sheet10")
            tenth_data=[]
            # 对ws的单个单元格传入数据
            data_excel = []
            # 将原始处理数据中的每对数据以列表形式传入data_excel列表
            for each in first_data:
                each = list(each)
                data_excel.append(each)
            # 将data_excel列表内的内容存入工作表
            for each in data_excel:
                ws.append(each)
                tenth_data.append(each)
            data_excel = []
            for each in second_data:
                each = list(each)
                data_excel.append(each)
            for each in data_excel:
                ws2.append(each)
                tenth_data.append(each)
            data_excel = []
            for each in third_data:
                each = list(each)
                data_excel.append(each)
            for each in data_excel:
                ws3.append(each)
                tenth_data.append(each)
            data_excel = []
            for each in fourth_data:
                each = list(each)
                data_excel.append(each)
            for each in data_excel:
                ws4.append(each)
                tenth_data.append(each)
            data_excel = []
            for each in fifth_data:
                each = list(each)
                data_excel.append(each)
            for each in data_excel:
                ws5.append(each)
                tenth_data.append(each)
            data_excel = []
            for each in sixth_data:
                each = list(each)
                data_excel.append(each)
            for each in data_excel:
                ws6.append(each)
                tenth_data.append(each)
            data_excel = []
            for each in seventh_data:
                each = list(each)
                data_excel.append(each)
            for each in data_excel:
                ws7.append(each)
                tenth_data.append(each)
            data_excel = []
            for each in eighth_data:
                each = list(each)
                if each[0]=="\n":
                    each[0]=str("换行符")
                    data_excel.append(each)
                elif each[0]==" ":
                    each[0]=str("空格")
                    data_excel.append(each)
                elif each[0]=="\xa0":
                    each[0]=str("不间断空白符")
                    data_excel.append(each)
                elif each[0]!="":
                    data_excel.append(each)
            for each in data_excel:
                ws8.append(each)
                tenth_data.append(each)
            data_excel = []
            for each in night_data:
                each = list(each)
                data_excel.append(each)
            for each in data_excel:
                ws9.append(each)
                tenth_data.append(each)
            tenth_data.sort(key=lambda x: x[1], reverse=True)
            for each in tenth_data:
                ws10.append(each)
            # 注意：上述两个append方法是意义完全不同的两个方法
            global i
            i = '处理历史.xlsx'
            try:
                wb.save('处理历史.xlsx')
                self.plainText.appendPlainText("文本处理完成")
            except PermissionError:
                self.plainText.appendPlainText("处理结果未保存，请关闭 处理历史.xlsx 后重试")

    def action_saveAs_triggered(self):  #保存到新文件中
        string = self.plainText.toPlainText()
        self.plainText.appendPlainText("正在制表")
        if string != "":
            name,fil=QFileDialog.getSaveFileName(self,"另存文件","d:\\","表格文件(*.xlsx)")
            if name != "":
                import matplotlib.pyplot as plt
                from wordcloud import WordCloud
                from openpyxl import Workbook  # Demo6_32.py
                from openpyxl.chart import Reference, BarChart, BarChart3D, label, PieChart, PieChart3D, DoughnutChart
                from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
                from openpyxl.chart.label import DataLabelList
                from openpyxl.chart.text import RichText
                from openpyxl.drawing.image import Image

                wbook = Workbook()
                wsheet = wbook.active
                score = [['全文字数',count_w],['查询词', "出现次数"]]
                for y in choose_data:
                    y=y.replace("  ",",")
                    a=y.split(",")
                    a[1]=int(a[1])
                    score.append(a)
                score2=score
                for item in score:
                    wsheet.append(item)
                bar1 = BarChart()  # 创建条形图对象
                bar3D = BarChart3D()  # 创建条形图对象
                col1 = BarChart()  # 创建条形图对象
                col3D = BarChart3D()  # 创建条形图对象
                bar1.type = bar3D.type = 'bar'
                col1.type = col3D.type = 'col'

                bar1.title = bar3D.title = "水平 Bar Chart"  # 设置名称
                col1.title = col3D.title = "竖直 Bar Chart"  # 设置名称

                bar1.style = bar3D.style = col1.style = col3D.style = 15  # 设置样式
                bar1.x_axis.title = col1.x_axis.title = bar3D.x_axis.title = col3D.x_axis.title = '日期'  # x轴名称
                bar1.y_axis.title = col1.y_axis.title = bar3D.y_axis.title = col3D.y_axis.title = '出现次数'  # y轴名称

                xLabel = Reference(wsheet, min_col=1, min_row=3, max_col=1, max_row=wsheet.max_row)  # 设置x轴坐标数据
                yData = Reference(wsheet, min_col=2, min_row=2, max_col=2, max_row=wsheet.max_row)  # 设置y轴数据

                bar1.add_data(yData, titles_from_data=True)  # 添加y轴数据，数据名称来自数据的第1个值
                bar3D.add_data(yData, titles_from_data=True)  # 添加y轴数据，数据名称来自数据的第1个值
                col1.add_data(yData, titles_from_data=True)  # 添加y轴数据，数据名称来自数据的第1个值
                col3D.add_data(yData, titles_from_data=True)  # 添加y轴数据，数据名称来自数据的第1个值

                col3D.series[0].shape = 'cylinder'  # 设置形状

                bar1.set_categories(xLabel)  # 添加x轴数据
                bar3D.set_categories(xLabel)  # 添加x轴数据
                col1.set_categories(xLabel)  # 添加x轴数据
                col3D.set_categories(xLabel)  # 添加x轴数据

                bar1.width = bar3D.width = col1.width = col3D.width = 30  # 设置高度
                bar1.height = bar3D.height = col1.height = col3D.height = 20  # 设置宽度

                bar1.dLbls = bar3D.dLbls = col1.dLbls = col3D.dLbls = label.DataLabelList()  # 设置数标
                bar1.dLbls.showVal = bar3D.dLbls.showVal = col1.dLbls.showVal = col3D.dLbls.showVal = True

                pie = PieChart()
                pie3D = PieChart3D()
                doughnut = DoughnutChart()

                pie.title = pie3D.title = doughnut.title = "查询词出现次数"

                label = Reference(wsheet, min_col=1, min_row=3, max_row=wsheet.max_row)
                score = Reference(wsheet, min_col=2, min_row=2, max_row=wsheet.max_row)

                pie.add_data(score, titles_from_data=True)
                pie3D.add_data(score, titles_from_data=True)
                doughnut.add_data(score, titles_from_data=True)

                pie.set_categories(label)
                pie3D.set_categories(label)
                doughnut.set_categories(label)

                pie.width = pie3D.width = doughnut.width = 30
                pie.height = pie3D.height = doughnut.height = 20

                pie.dLbls = pie3D.dLbls = doughnut.dLbls = DataLabelList()
                pie.dLbls.showCatName = pie3D.dLbls.showCatName = doughnut.dLbls.showCatName = True  # 标签显示
                pie.dLbls.showPercent = pie3D.dLbls.showPercent = doughnut.dLbls.showPercent = True  # 百分比显示

                score2.remove(['查询词', '出现次数'])
                score2.remove(["全文字数",count_w])
                score2=dict(score2)
                wc = WordCloud(
                    scale=12,
                    background_color="white",
                    font_path="C:/Windows/Fonts/simfang.ttf",
                    max_words=120,
                    width=1000,
                    height=600,
                )
                word_cloud = wc.generate_from_frequencies(score2)
                # 写词云图片
                word_cloud.to_file("create_word_cloud_by_words_count.jpg")

                img=Image("create_word_cloud_by_words_count.jpg")
                new_size=(600,360)
                img.width,img.height=new_size
                wsheet.add_image(img, "D121")
                wsheet.add_chart(bar1, "D1")  # 图表添加进工作表格中
                wsheet.add_chart(bar3D, "V1")  # 图表添加进工作表格中
                wsheet.add_chart(col1, "D41")  # 图表添加进工作表格中
                wsheet.add_chart(col3D, "V41")  # 图表添加进工作表格中
                wsheet.add_chart(pie, "D81")
                wsheet.add_chart(pie3D, "V81")
                wsheet.add_chart(doughnut, "AN81")
                try:
                    wbook.save(r"{}".format(name))
                    self.plainText.appendPlainText("处理结果已保存")
                except PermissionError:
                    self.plainText.appendPlainText("处理结果未保存，请关闭选择的表格文件后重试")

    def action_choose_triggered(self):
        self.child_window =Child()
        self.child_window.show()

    def action_help_triggered(self):
        self.child_window =Child2()
        self.child_window.show()

class Child(QWidget):
    def __init__(self,parent=None):
        super().__init__(parent)
        self.showMaximized()
        self.fileName =r"{}".format(i)
        self.reference_Model=QStringListModel(self)  #从Excel中读取数据后，存储数据的模型
        self.selection_Model=QStringListModel(self)  #选择数据后，存储选择数据的模型
        self.setup_Ui()  #建立界面
        self.data_import()  #从Excel中读取数据
        self.view_clicked()  #单击视图控件，判断按钮是否激活或失效
    def setup_Ui(self):  #建立界面
        self.setStyleSheet("QWidget{background-color:rgb(212,202,210)}")
        menuBar = QToolBar(self)  # 定义菜单栏
        pos_menu = menuBar  # 定义文件菜单
        pos_menu.setStyleSheet("background-color:rgb(212,241,118)")
        pos_menu.addSeparator()
        all_name = pos_menu.addAction("全部")
        qnr_name = pos_menu.addAction("人名")
        qns_name = pos_menu.addAction("地名")
        qnt_name = pos_menu.addAction("机构名")
        qn_name = pos_menu.addAction("形容词")
        qa_name = pos_menu.addAction("物名")
        qv_name = pos_menu.addAction("动名词")
        qvn_name = pos_menu.addAction("纯动词")
        number_name = pos_menu.addAction("数")
        other_name = pos_menu.addAction(("其他"))

        label1 = QLabel("待选词")
        self.listView_1=QListView()  #列表视图控件，显示Excel中的数据的控件
        v1=QVBoxLayout()
        v1.addWidget(label1)
        v1.addWidget(self.listView_1)
        v1.addWidget(menuBar)
        label2=QLabel("需处理词")
        self.listView_2=QListView()  #列表视图控件，显示选中的数据
        self.btn_add = QPushButton("添加")
        self.btn_insert = QPushButton("插入")
        self.btn_delete = QPushButton("删除")
        self.btn_done = QPushButton("确定")
        self.btn_add.setStyleSheet("background-color:rgb(180,226,216)")
        self.btn_insert.setStyleSheet("background-color:rgb(222,243,211)")
        self.btn_done.setStyleSheet("background-color:rgb(213,243,236)")
        self.btn_delete.setStyleSheet("background-color:rgb(203,203,214)")
        self.listView_1.setStyleSheet("background-color:rgb(206,225,225)")
        self.listView_2.setStyleSheet("background-color:rgb(208,222,216)")
        h1=QHBoxLayout()
        h1.addWidget(self.btn_add)
        h1.addWidget(self.btn_insert)
        h1.addWidget(self.btn_delete)
        h1.addWidget(self.btn_done)
        v2=QVBoxLayout()
        v2.addWidget(label2)
        v2.addWidget(self.listView_2)
        v2.addLayout(h1)
        h2=QHBoxLayout(self)
        h2.addLayout(v1)
        h2.addLayout(v2)

        self.listView_1.setModel(self.reference_Model)  #设置模型
        self.listView_2.setModel(self.selection_Model)  #设置模型
        self.listView_1.setSelectionMode(QListView.ExtendedSelection) #设置选择模式
        self.listView_2.setSelectionMode(QListView.ExtendedSelection) #设置选择模式

        self.btn_add.clicked.connect(self.btn_add_clicked)
        self.btn_insert.clicked.connect(self.btn_insert_clicked)
        self.btn_delete.clicked.connect(self.btn_delete_clicked)
        self.btn_done.clicked.connect(self.btn_done_clicked)
        self.listView_1.clicked.connect(self.view_clicked)
        self.listView_2.clicked.connect(self.view_clicked)

        qnr_name.triggered.connect(self.qnr_name_triggered)
        qns_name.triggered.connect(self.qns_name_triggered)
        qnt_name.triggered.connect(self.qnt_name_triggered)
        qn_name.triggered.connect(self.qn_name_triggered)
        qa_name.triggered.connect(self.qa_name_triggered)
        qv_name.triggered.connect(self.qv_name_triggered)
        qvn_name.triggered.connect(self.qvn_name_triggered)
        other_name.triggered.connect(self.other_name_triggered)
        number_name.triggered.connect(self.number_name_triggered)
        all_name.triggered.connect(self.all_name_triggered)

        self.listView_1.setEditTriggers(QAbstractItemView.NoEditTriggers)#设置内容不可编辑
        self.listView_2.setEditTriggers(QAbstractItemView.NoEditTriggers)#设置内容不可编辑

    def qnr_name_triggered(self):
        global sheet_name
        sheet_name="sheet1"
        self.data_import()
        self.view_clicked()
    def qns_name_triggered(self):
        global sheet_name
        sheet_name = "sheet2"
        self.data_import()
        self.view_clicked()
    def qnt_name_triggered(self):
        global sheet_name
        sheet_name = "sheet3"
        self.data_import()
        self.view_clicked()
    def qn_name_triggered(self):
        global sheet_name
        sheet_name = "sheet5"
        self.data_import()
        self.view_clicked()
    def qa_name_triggered(self):
        global sheet_name
        sheet_name = "sheet4"
        self.data_import()
        self.view_clicked()
    def qv_name_triggered(self):
        global sheet_name
        sheet_name = "sheet7"
        self.data_import()
        self.view_clicked()
    def qvn_name_triggered(self):
        global sheet_name
        sheet_name = "sheet6"
        self.data_import()
        self.view_clicked()
    def other_name_triggered(self):
        global sheet_name
        sheet_name = "sheet8"
        self.data_import()
        self.view_clicked()
    def number_name_triggered(self):
        global sheet_name
        sheet_name = "sheet9"
        self.data_import()
        self.view_clicked()
    def all_name_triggered(self):
        global sheet_name
        sheet_name = "sheet10"
        self.data_import()
        self.view_clicked()

    def data_import(self):
        if os.path.exists(self.fileName):
            wbook = load_workbook(self.fileName)
            if "sheet1" in wbook.sheetnames:
                global student
                wsheet = wbook[sheet_name]
                cell_range = wsheet[wsheet.dimensions]  #获取Excel中数据存储的范围
                student=list()
                for cell_row in cell_range:  # cell_row是Excel行单元格元组
                    string = ""
                    for cell in cell_row:
                        string=string+str(cell.value)+"  "  #获取Excel单元格中的数据
                    student.append(string.strip())
                self.reference_Model.setStringList(student)  #在模型中添加数据列表
                self.selection_Model.setStringList(choose_data)

    def btn_add_clicked(self):  #添加按钮的槽函数
        global choose_data
        a=len(self.listView_1.selectedIndexes())
        b=0
        while a:
            selectedIndexes=self.listView_1.selectedIndexes()
            index= selectedIndexes[b]
            b=b+1
            string = self.reference_Model.data(index,Qt.DisplayRole)  #获取数据
            a=a-1
            count = self.selection_Model.rowCount()  #获取行的数量
            self.selection_Model.insertRow(count)  #在末尾插入数据
            last_index = self.selection_Model.index(count, 0, QModelIndex()) #获取末尾的索引
            self.selection_Model.setData(last_index,string,Qt.DisplayRole) #设置末尾的数据
            choose_data.append(string)
        self.view_clicked()  #控制按钮的激活与失效
    def btn_insert_clicked(self):  #插入按钮的槽函数
        global choose_data
        a=len(self.listView_1.selectedIndexes())
        b=0
        while a:
            selectedIndexs_1 = self.listView_1.selectedIndexes()  # 获取选中数据项的索引
            selectedIndex_2 = self.listView_2.selectedIndexes()  # 获取选中数据项的索引
            index= selectedIndexs_1[b]
            b=b+1
            string = self.reference_Model.data(index, Qt.DisplayRole)
            a=a-1
            row=selectedIndex_2[0].row()
            self.selection_Model.insertRow(row)
            index=self.selection_Model.index(row)
            self.selection_Model.setData(index, string, Qt.DisplayRole)
            choose_data.insert(row,string)
        self.view_clicked()
    def btn_delete_clicked(self):  #删除按钮的槽函数
        global choose_data
        while len(self.listView_2.selectedIndexes()):
            selectedIndexes=self.listView_2.selectedIndexes()
            index= selectedIndexes[0]
            string = self.selection_Model.data(index, Qt.DisplayRole)
            self.selection_Model.removeRow(index.row(), QModelIndex())
            choose_data.remove(string)
        self.view_clicked()
    def btn_done_clicked(self):
        self.close()

    def view_clicked(self):  #单击视图控件的槽函数，用于按钮的激活或失效
        n1=len(self.listView_1.selectedIndexes())  #获取选中数据项的数量
        n2 = len(self.listView_2.selectedIndexes())  #获取选中数据项的数量
        self.btn_add.setEnabled(n1)
        self.btn_insert.setEnabled(n1 and n2==1)
        self.btn_delete.setEnabled(n2)

class Child2(QWidget):
    def __init__(self,parent=None):
        super().__init__(parent)
        self.showMaximized()
        self.widget2_setupUi()

    def widget2_setupUi(self):  # 建立主程序界面
        self.setStyleSheet("QWidget{background-color:rgb(240,230,199)}")
        self.plainText = QPlainTextEdit(self)  # 显示数据控件
        self.plainText.setReadOnly(True)
        font = QFont()
        font.setFamily("SimSun")  # 宋体
        font.setBold(True)
        font.setPointSizeF(16)
        self.plainText.setFont(font)
        v2= QVBoxLayout(self)  #主程序界面的布局
        v2.addWidget(self.plainText)
        self.plainText.appendPlainText('此软件为贵州师范大学附属中学研究性学习0202小组 吴张元和肖自强同学寒假项目')
        self.plainText.appendPlainText('指引:')
        self.plainText.appendPlainText('1,点击 文件 ')
        self.plainText.appendPlainText('2,点击 打开 ,选择要处理的文本 ')
        self.plainText.appendPlainText('3,点击 选择 ,选择要处理的词汇 ')
        self.plainText.appendPlainText('4,点击 保存处理文件至 ')
        self.plainText.appendPlainText('5,退出程序,打开选择的表格文件，收获')



if __name__ == '__main__':
    app=QApplication(sys.argv)
    window = myWindow()
    window.show()
    sys.exit(app.exec())
